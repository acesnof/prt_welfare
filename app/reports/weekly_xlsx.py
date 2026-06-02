import os
from copy import copy
from openpyxl import Workbook, load_workbook

MONTH_ABBR_EN = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

DATE_CELLS = ["D4", "G4", "J4", "M4", "P4", "S4", "V4"]
DAY_START_COLS = [4, 7, 10, 13, 16, 19, 22]  # D, G, J, M, P, S, V


def _procurar_template(docs_dir):
    candidatos = [
        os.path.join(docs_dir, "meals_request_weekly.xlsx"),
        os.path.join(docs_dir, "meals_request_weekly.xlsm"),
        os.path.join(docs_dir, "meals_request_weekly"),
    ]
    for caminho in candidatos:
        if os.path.exists(caminho):
            return caminho
    return None


def _formatar_data_en(data_ref):
    return f"{data_ref.day}-{MONTH_ABBR_EN[data_ref.month - 1]}"


def _criar_workbook_fallback():
    wb = Workbook()
    ws = wb.active
    ws.title = "Meals Request"
    ws["A1"] = "PT CN - Weekly Meals Request"
    ws["A3"] = "Date"
    ws["A7"] = "Totals"
    for idx, start_col in enumerate(DAY_START_COLS):
        ws.cell(row=6, column=start_col, value="Breakfast")
        ws.cell(row=6, column=start_col + 1, value="Lunch")
        ws.cell(row=6, column=start_col + 2, value="Dinner")
    return wb


def gerar_meals_request_weekly(docs_dir, destino, totais_semana):
    """
    Preenche o modelo docs/meals_request_weekly(.xlsx) com totais semanais.

    totais_semana deve ser uma lista com 7 dicionários:
    {
        "data": date,
        "pequeno_almoco": int,
        "almoco": int,
        "jantar": int,
    }
    """
    template = _procurar_template(docs_dir)

    if template:
        wb = load_workbook(template)
        ws = wb.active
    else:
        wb = _criar_workbook_fallback()
        ws = wb.active

    for idx, info in enumerate(totais_semana[:7]):
        data_ref = info["data"]
        ws[DATE_CELLS[idx]] = _formatar_data_en(data_ref)

        col = DAY_START_COLS[idx]
        ws.cell(row=7, column=col, value=int(info.get("pequeno_almoco") or 0))
        ws.cell(row=7, column=col + 1, value=int(info.get("almoco") or 0))
        ws.cell(row=7, column=col + 2, value=int(info.get("jantar") or 0))

    os.makedirs(os.path.dirname(destino), exist_ok=True)
    wb.save(destino)
    return destino
