import os
import re
import calendar
from copy import copy
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell

MONTH_ABBR_EN = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

SHEETS = ["F1", "F2", "F3", "F4", "F5"]
START_ROW = 9
MAX_PER_SHEET = 15

POSTO_ORDEM = {
    "OF-6": 1, "OF-06": 1,
    "OF-5": 2, "OF-05": 2,
    "OF-4": 3, "OF-04": 3,
    "OF-3": 4, "OF-03": 4,
    "OF-2": 5, "OF-02": 5,
    "OF-1": 6, "OF-01": 6,
    "OR-9": 7, "OR-09": 7,
    "OR-8": 8, "OR-08": 8,
    "OR-7": 9, "OR-07": 9,
    "OR-6": 10, "OR-06": 10,
    "OR-5": 11, "OR-05": 11,
    "OR-4": 12, "OR-04": 12,
    "OR-3": 13, "OR-03": 13,
    "OR-2": 14, "OR-02": 14,
    "OR-1": 15, "OR-01": 15,
}


def _posto_key(posto):
    """Ordenação militar: OF-6..OF-1, depois OR-9..OR-1.
    Aceita OF-06/OF6/OF 6, OR-05/OR5, etc.
    """
    bruto = (posto or "").strip().upper()
    if bruto in POSTO_ORDEM:
        return POSTO_ORDEM[bruto]

    m = re.search(r"(OF|OR)\s*-?\s*0?(\d+)", bruto)
    if not m:
        return 99

    grupo = m.group(1)
    numero = int(m.group(2))

    if grupo == "OF":
        # OF-6 é o mais antigo, OF-1 o menos antigo dentro dos OF.
        return 7 - numero if 1 <= numero <= 6 else 99

    if grupo == "OR":
        # OR-9 é o mais antigo, OR-1 o menos antigo dentro dos OR.
        return 7 + (9 - numero) if 1 <= numero <= 9 else 99

    return 99


def _antiguidade_key(valor):
    valor = (valor or "").strip()
    if not valor:
        return "9999-12-31"

    # ISO datetime/date: 2026-05-01 or 2026-05-01 08:00
    iso = valor[:10]
    if len(iso) == 10 and iso[4] == "-" and iso[7] == "-":
        yyyy, mm, dd = iso.split("-")
        return f"{yyyy.zfill(4)}-{mm.zfill(2)}-{dd.zfill(2)}"

    # dd/mm/yyyy, dd.mm.yyyy, dd-mm-yyyy, with optional time after it.
    data_txt = valor.split()[0]
    for sep in ("/", ".", "-"):
        partes = data_txt.split(sep)
        if len(partes) == 3:
            dd, mm, yyyy = partes
            if len(yyyy) == 4:
                return f"{yyyy.zfill(4)}-{mm.zfill(2)}-{dd.zfill(2)}"

    return "9999-12-31"


def _normalizar_posto_assinatura(posto):
    bruto = (posto or "").strip().upper()
    m = re.search(r"(OF|OR)\s*-?\s*0?(\d+)", bruto)
    if not m:
        return bruto
    return f"{m.group(1)}-{int(m.group(2))}"

def _sort_key_linha(linha):
    return (
        _posto_key(linha.get("posto")),
        _antiguidade_key(linha.get("antiguidade")),
        (linha.get("sobrenome") or "").upper(),
        (linha.get("nome") or "").upper(),
    )


def _procurar_template(docs_dir):
    candidatos = [
        os.path.join(docs_dir, "meals_reimbursment.xlsx"),
        os.path.join(docs_dir, "meals_reimbursement.xlsx"),
        os.path.join(docs_dir, "meals_reimbursment"),
        os.path.join(docs_dir, "meals_reimbursement"),
    ]
    for caminho in candidatos:
        if os.path.exists(caminho):
            return caminho
    return None


def _formatar_inicio_fim(ano, mes):
    ultimo = calendar.monthrange(ano, mes)[1]
    return f"01.{mes:02d}.{ano}", f"{ultimo:02d}.{mes:02d}.{ano}"


def _formatar_data_exportacao():
    hoje = date.today()
    return f"{hoje.day:02d}/{MONTH_ABBR_EN[hoje.month - 1].lower()}/{hoje.year}"



def _celula_escrita(ws, cell_ref):
    cell = ws[cell_ref]
    if not isinstance(cell, MergedCell):
        return cell

    for merged_range in ws.merged_cells.ranges:
        if cell_ref in merged_range:
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col)

    return cell


def _set_value(ws, cell_ref, value):
    _celula_escrita(ws, cell_ref).value = value


def _cell_escrita_por_pos(ws, row, column):
    cell = ws.cell(row=row, column=column)
    if not isinstance(cell, MergedCell):
        return cell

    coord = cell.coordinate
    for merged_range in ws.merged_cells.ranges:
        if coord in merged_range:
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col)

    return cell

def _copiar_estilo(origem, destino):
    if origem.has_style:
        destino.font = copy(origem.font)
        destino.fill = copy(origem.fill)
        destino.border = copy(origem.border)
        destino.alignment = copy(origem.alignment)
        destino.number_format = origem.number_format
        destino.protection = copy(origem.protection)


def _criar_workbook_fallback():
    wb = Workbook()
    ws = wb.active
    ws.title = "F1"
    for nome in SHEETS[1:]:
        wb.create_sheet(nome)

    for ws in wb.worksheets:
        ws["A1"] = "MEALS REIMBURSEMENT"
        ws["B4"] = "Start"
        ws["C4"] = "End"
        ws["A8"] = "Rank"
        ws["B8"] = "Surname"
        ws["C8"] = "Name"
        ws["D8"] = "Welfare"
        ws["E8"] = "Cohesion"
        ws["F8"] = "Reimbursement"
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 16
    return wb


def _garantir_folhas(wb):
    # Se o modelo não trouxer F1..F5, cria as folhas em falta.
    primeira = wb[wb.sheetnames[0]]
    for nome in SHEETS:
        if nome not in wb.sheetnames:
            nova = wb.copy_worksheet(primeira)
            nova.title = nome


def _limpar_linhas(ws):
    limpas = set()
    for row in range(START_ROW, START_ROW + MAX_PER_SHEET):
        for col in range(1, 7):
            cell = _cell_escrita_por_pos(ws, row, col)
            coord = cell.coordinate
            if coord not in limpas:
                cell.value = None
                limpas.add(coord)


def _preencher_cabecalhos(ws, ano, mes, data_inicio_override=None):
    inicio, fim = _formatar_inicio_fim(ano, mes)
    _set_value(ws, "B5", data_inicio_override or inicio)
    _set_value(ws, "C5", fim)
    _set_value(ws, "B30", _formatar_data_exportacao())


def _preencher_assinatura(ws, primeiro):
    if primeiro:
        posto = _normalizar_posto_assinatura(primeiro.get("posto", ""))
        sobrenome = (primeiro.get("sobrenome", "") or "").strip().upper()
        _set_value(ws, "B31", f"{posto} {sobrenome}".strip())
    else:
        _set_value(ws, "B31", "")


def gerar_reembolso_mensal(docs_dir, destino, ano, mes, linhas, senior_assinatura=None, data_inicio_override=None):
    """
    Preenche docs/meals_reimbursment.xlsx com os reembolsos mensais.

    Mapeamento por folha F1..F5:
    - B5/C5: início/fim do mês.
    - A9:F23: 15 militares por folha.
    - B30: data da exportação.
    - B31: posto + sobrenome do militar mais antigo da lista exportada.
    """
    # Garante que B31 e todas as folhas respeitam a antiguidade:
    # OF antes de OR; dentro do mesmo grupo, posto mais alto primeiro;
    # se o posto for igual, antiguidade mais antiga primeiro.
    linhas = sorted(linhas, key=_sort_key_linha)

    template = _procurar_template(docs_dir)

    if template:
        wb = load_workbook(template)
        _garantir_folhas(wb)
    else:
        wb = _criar_workbook_fallback()

    primeiro = senior_assinatura

    for sheet_idx, sheet_name in enumerate(SHEETS):
        ws = wb[sheet_name]
        _preencher_cabecalhos(ws, ano, mes, data_inicio_override=data_inicio_override)
        _preencher_assinatura(ws, primeiro)
        _limpar_linhas(ws)

        bloco = linhas[sheet_idx * MAX_PER_SHEET:(sheet_idx + 1) * MAX_PER_SHEET]
        total_reembolso_folha = 0

        for idx, linha in enumerate(bloco):
            row = START_ROW + idx
            reembolso = int(linha.get("reimbursement") or 0)
            total_reembolso_folha += reembolso

            valores = [
                linha.get("posto", ""),
                linha.get("sobrenome", ""),
                linha.get("nome", ""),
                int(linha.get("welfare") or 0),
                int(linha.get("cohesion") or 0),
                reembolso,
            ]
            for col, valor in enumerate(valores, start=1):
                cell = _cell_escrita_por_pos(ws, row, col)
                cell.value = valor
                # Mantém o estilo da linha seguinte/modelo quando possível.
                if row > START_ROW:
                    origem = _cell_escrita_por_pos(ws, START_ROW, col)
                    _copiar_estilo(origem, cell)

        # F24: soma total do valor dos reembolsos dos militares desta folha.
        _set_value(ws, "F24", total_reembolso_folha)

    os.makedirs(os.path.dirname(destino), exist_ok=True)
    wb.save(destino)
    return destino
